import scrapy
import time
from scrapy_selenium import SeleniumRequest
from scrapy.selector import Selector
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException
from openpyxl import load_workbook

wb = load_workbook('BD_CADASTRO_NUMERADO_AGO_TESTE.xlsx')
ws, ws2 = wb['Fundos'], wb['Fundos_Cota']
lista_cnpj = []

for row in ws.iter_rows(values_only=True):
    cnpj = row[1]
    lista_cnpj.append(cnpj)

for row in ws2.iter_rows(values_only=True):
    cnpj = row[1] 
    lista_cnpj.append(cnpj)

ignored_exceptions=(StaleElementReferenceException, NoSuchElementException)

class CvmSeleniumSpider(scrapy.Spider):
    name = "cvm_selenium"
    # allowed_domains = ["cvmweb.cvm.gov.br"]
    # start_urls = ["https://cvmweb.cvm.gov.br/swb/default.asp?sg_sistema=fundosreg"]

    def start_requests(self):
        yield SeleniumRequest(
            url='https://cvmweb.cvm.gov.br/swb/default.asp?sg_sistema=fundosreg',
            wait_time=3,
            callback=self.parse,
        )

    def parse(self, response):
        cnpj = lista_cnpj[0]
        driver = response.meta['driver']

        #pesquisa e troca para o frame correto sem precisar declarar uma variável nova no processo
        WebDriverWait(driver, 10).until(
        EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
        )
        

        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='txtCNPJNome']"))
        )
        search_input.clear()
        search_input.send_keys(cnpj)
        search_input.send_keys(Keys.ENTER)

        find_name_to_click = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[@id='ddlFundos__ctl0_lnkbtn1']"))
        )
        find_name_to_click.click()

        driver.switch_to.default_content()

        #pesquisa e troca para o frame correto sem precisar declarar uma variável nova no processo
        WebDriverWait(driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
        )

        dados_diarios = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[@id='Hyperlink2']"))
        )
        dados_diarios.click()

        driver.switch_to.default_content()

        #pop-up com mensagem irrelevante aparece e precisa ser fechado.
        try:
            alert = WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert.accept()
        except:
            pass

        #pesquisa e troca para o frame correto sem precisar declarar uma variável nova no processo
        WebDriverWait(driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
        )

        select_data_pesquisa = Select(WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//select[@name='ddComptc']"))
        ))

        data_pesquisa_to_iterate = select_data_pesquisa.options

        for data_pesquisa in data_pesquisa_to_iterate:
            time.sleep(2)
            
            is_stale = True
            while is_stale:
                is_stale, linhas_tabela = self.find_data_pesquisa(driver, data_pesquisa)
                


            #se não tiver valor na quota (coluna 2), já filtra e retira a linha
            linhas_validas = [linha for linha in linhas_tabela if linha.find_element(By.XPATH, "td[2]").text.strip()] 
            dados_diarios_list = []
            if linhas_validas:
                ultima_linha = linhas_validas[-1]
                dados = ultima_linha.find_elements(By.XPATH, "td")
                dados_diarios = {
                    "Dia": dados[0].text.strip(),
                    "Quota": dados[1].text.strip(),
                    #"Captação no Dia": dados[2].text.strip(),
                    #"Resgate no Dia": dados[3].text.strip(),
                    "Patrimônio Líquido": dados[4].text.strip(),
                    #"Total da Carteira": dados[5].text.strip(),
                    "Número de Cotistas": dados[6].text.strip(),
                    #"Data da Próxima Informação do PL": dados[7].text.strip(),
                }
                dados_diarios_list.append(dados_diarios)

                yield {
                    #'url': driver.current_url,
                    #'fundo': response_obj.get(),
                    'dados_diarios': dados_diarios
                }

    def find_data_pesquisa(self, driver, data_pesquisa):
        try:
            # Aguardar mais um pouco antes de continuar, caso a página ainda esteja atualizando
            time.sleep(2)  # Atraso adicional
            
            select_data_pesquisa = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//select[@name='ddComptc']"))
            )

            select_data_pesquisa = Select(select_data_pesquisa)
            select_data_pesquisa.select_by_value(data_pesquisa.text)

            driver.switch_to.default_content()

            #pesquisa e troca para o frame correto sem precisar declarar uma variável nova no processo
            WebDriverWait(driver, 10).until(
                EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
            )

            #ignora a primeira linha, >1 (pula o primeiro tr)
            linhas_tabela = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//table[@id='dgDocDiario']//tr[position()>1]"))
            )
        except StaleElementReferenceException:
            return True, None
        return False, linhas_tabela