import scrapy
import time
from scrapy_selenium import SeleniumRequest
from scrapy.selector import Selector
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException
from openpyxl import load_workbook

wb = load_workbook('BD_CADASTRO_NUMERADO_AGO_TESTE.xlsx')
ws, ws2 = wb['Fundos'], wb['Fundos_Cota']
lista_cnpj = []

for row in ws.iter_rows(values_only=True):
    cnpj = row[1]
    lista_cnpj.append(cnpj)

for row in ws2.iter_rows(values_only=True):
    cnpj = row[1] 
    lista_cnpj.append(cnpj)


class CvmSeleniumSpider(scrapy.Spider):
    name = "cvm_step"
    # allowed_domains = ["cvmweb.cvm.gov.br"]
    # start_urls = ["https://cvmweb.cvm.gov.br/swb/default.asp?sg_sistema=fundosreg"]

    def start_requests(self):
        yield SeleniumRequest(
            url='https://cvmweb.cvm.gov.br/swb/default.asp?sg_sistema=fundosreg',
            wait_time=3,
            callback=self.parse,
        )

    def parse(self, response):
        driver = response.request.meta['driver']

        #pesquisa e troca para o frame correto sem precisar declarar uma variável nova no processo
        WebDriverWait(driver, 10).until(
        EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
        )

        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='txtCNPJNome']"))
        )
        search_input.clear()
        search_input.send_keys(lista_cnpj[0])
        search_input.send_keys(Keys.ENTER)

        find_name_to_store = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[@id='ddlFundos__ctl0_lnkbtn1']"))
        )
        stored_name = find_name_to_store.text
        find_name_to_store.click()

        driver.switch_to.default_content()

        #pesquisa e troca para o frame correto sem precisar declarar uma variável nova no processo
        WebDriverWait(driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
        )

        dados_diarios = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[@id='Hyperlink2']"))
        )
        dados_diarios.click()

        driver.switch_to.default_content()

        #pop-up com mensagem irrelevante aparece e precisa ser fechado.
        try:
            alert = WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert.accept()
        except:
            pass

        WebDriverWait(driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
        )

        dropdown_data_pesquisa = Select(WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//select[@name='ddComptc']"))
        ))

        dados_diarios_list = []
        for i, _ in enumerate(dropdown_data_pesquisa.options):
            driver.switch_to.default_content()

            WebDriverWait(driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//frame[@name='Main']"))
            )

            dropdown_data_pesquisa = Select(WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//select[@name='ddComptc']"))
            ))
            dropdown_data_pesquisa.select_by_index(i)

            name_selected_dropdown = Select(WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//select[@name='ddComptc']"))
            ))
            name_selected_dropdown = name_selected_dropdown.first_selected_option.get_attribute("value")
            
            linhas_tabela = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//table[@id='dgDocDiario']//tr[position()>1]"))
            )
            linhas_validas = [linha for linha in linhas_tabela if linha.find_element(By.XPATH, "td[2]").text.strip()] 
            
            if linhas_validas:
                ultima_linha = linhas_validas[-1]
                dados = ultima_linha.find_elements(By.XPATH, "td")
                dados_diarios = {
                    "Mês": name_selected_dropdown,
                    "Dia": dados[0].text.strip(),
                    "Quota": dados[1].text.strip(),
                    #"Captação no Dia": dados[2].text.strip(),
                    #"Resgate no Dia": dados[3].text.strip(),
                    "Patrimônio Líquido": dados[4].text.strip(),
                    #"Total da Carteira": dados[5].text.strip(),
                    "Número de Cotistas": dados[6].text.strip(),
                    #"Data da Próxima Informação do PL": dados[7].text.strip(),
                }
                dados_diarios_list.append(dados_diarios)

        yield {
            'name': stored_name,
            'dados': dados_diarios
        }